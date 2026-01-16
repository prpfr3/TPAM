import openpyxl
from datetime import date
from django.core.management.base import BaseCommand
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from locations.models import Location, Route, ELR, EventType, LocationHistoricEvent


class Command(BaseCommand):
    help = "Uploads historic events from an Excel file using object slugs with CLI command of" \
    "python manage.py Location_Events_Load 'pythoC:\Users\paulf\OneDrive\Data\TPAM\Location_Events_Upload.xlsx'"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")

    def parse_custom_date(self, date_str):
        """
        Converts DD-MM-YYYY or ??-MM-YYYY or ??-??-YYYY to a valid date object.
        Replaces '??' with '01' to satisfy the DateField.
        """
        if not date_str:
            return None

        # Standardize: replace ? with 0
        clean_str = str(date_str).replace("?", "0").strip()
        parts = clean_str.split("-")

        if len(parts) != 3:
            return None

        day = int(parts[0]) if parts[0] != "00" else 1
        month = int(parts[1]) if parts[1] != "00" else 1
        year = int(parts[2])

        return date(year, month, day)

    def handle(self, *args, **options):
        file_path = options["file_path"]
        # file_path = r"C:\Users\paulf\OneDrive\Data\TPAM\Location Events Upload.xlsx"  # Hardcoded for testing

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Option A: Look for a specific name
            if "Events" in wb.sheetnames:
                sheet = wb["Events"]
            else:
                # Option B: Fallback to the very first sheet in the workbook
                sheet = wb.worksheets[0]
                self.stdout.write(
                    self.style.WARNING(
                        f"'Events' sheet not found. Using '{sheet.title}' instead."
                    )
                )

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Could not open file: {e}"))
            return

        # Prepare ContentTypes for lookup
        ct_map = {
            "location": ContentType.objects.get_for_model(Location),
            "route": ContentType.objects.get_for_model(Route),
            "elr": ContentType.objects.get_for_model(ELR),
        }

        # Cache EventTypes to avoid repetitive queries
        event_types = {et.code: et for et in EventType.objects.all()}

        count = 0
        # Assuming header row: Slug | ModelType | EventCode | DateStr | Description
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        try:
            with transaction.atomic():
                for row in rows:
                    obj_slug, model_type, event_code, date_str, desc = row

                    if not obj_slug or not model_type:
                        continue

                    # 1. Get ContentType and Target Object
                    ct = ct_map.get(model_type.lower())
                    if not ct:
                        self.stdout.write(
                            self.style.ERROR(f"Unknown model type: {model_type}")
                        )
                        continue

                    target_model = ct.model_class()
                    try:
                        target_obj = target_model.objects.get(slug=obj_slug)
                    except target_model.DoesNotExist:
                        self.stdout.write(
                            self.style.WARNING(
                                f"Slug {obj_slug} not found in {model_type}"
                            )
                        )
                        continue

                    # 2. Get Event Type
                    etype = event_types.get(event_code)
                    if not etype:
                        self.stdout.write(
                            self.style.WARNING(f"Event Code {event_code} not found")
                        )
                        continue

                    # 3. Parse Date
                    valid_date = self.parse_custom_date(date_str)
                    if not valid_date:
                        self.stdout.write(
                            self.style.ERROR(
                                f"Invalid date format: {date_str} for {obj_slug}"
                            )
                        )
                        continue

                    # 4. Create Event
                    LocationHistoricEvent.objects.create(
                        event_type=etype,
                        content_type=ct,
                        object_id=target_obj.id,
                        datefield=valid_date,
                        displaydate=date_str,
                        description=desc or "",
                    )
                    count += 1

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Transaction aborted. Error: {e}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Successfully uploaded {count} events."))
